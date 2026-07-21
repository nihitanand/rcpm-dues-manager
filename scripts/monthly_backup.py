#!/usr/bin/env python3
"""
RCPM Dues Manager — Monthly Backup Script
Fetches all data from Firebase Realtime Database, generates JSON + Excel backups,
and emails them via Gmail SMTP.

Run by GitHub Actions on a monthly cron schedule.
"""

import os
import json
import smtplib
import requests
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Environment / Secrets ──
FIREBASE_DB_URL     = os.environ["FIREBASE_DB_URL"].rstrip("/")
GMAIL_ADDRESS        = os.environ["GMAIL_ADDRESS"]
GMAIL_APP_PASSWORD   = os.environ["GMAIL_APP_PASSWORD"]
BACKUP_RECIPIENTS    = os.environ["BACKUP_RECIPIENTS"].split(",")

DB_PATH   = "rcpm/v1"
NOW       = datetime.now()
MONTH_STR = NOW.strftime("%B %Y")          # e.g. "July 2026"
FILE_TAG  = NOW.strftime("%Y-%m")          # e.g. "2026-07"

OUT_DIR = "/tmp/rcpm_backup"
os.makedirs(OUT_DIR, exist_ok=True)


def fetch_firebase_data():
    """Fetch the full rcpm/v1 node from Firebase Realtime Database (public read)."""
    url = f"{FIREBASE_DB_URL}/{DB_PATH}.json"
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    if not data:
        raise RuntimeError("No data returned from Firebase — check DB URL / rules.")
    return data


def to_list(value):
    """Firebase often returns dict-of-dicts instead of arrays — normalize to a list."""
    if not value:
        return []
    if isinstance(value, list):
        return [v for v in value if v]
    if isinstance(value, dict):
        return [v for v in value.values() if v]
    return []


def save_json_backup(data):
    path = os.path.join(OUT_DIR, f"RCPM_Backup_{FILE_TAG}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    return path


def save_excel_backup(data):
    members  = to_list(data.get("members"))
    payments = to_list(data.get("payments"))
    settings = data.get("settings", {}) or {}

    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    TITLE_FONT  = Font(bold=True, size=14, color="1B3A6B")

    # ── Sheet 1: Members ──
    ws1 = wb.active
    ws1.title = "Members"
    ws1["A1"] = "RCPM Dues Manager — Member Directory Backup"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:H1")

    headers = ["SN", "Primary Name", "Spouse Name", "Type", "Phone 1", "Phone 2", "ID"]
    for col, h in enumerate(headers, start=1):
        c = ws1.cell(row=3, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    row = 4
    for m in sorted(members, key=lambda x: x.get("sn", 0)):
        ws1.cell(row=row, column=1, value=m.get("sn"))
        ws1.cell(row=row, column=2, value=m.get("primary", ""))
        ws1.cell(row=row, column=3, value=m.get("spouse", ""))
        ws1.cell(row=row, column=4, value=m.get("type", ""))
        ws1.cell(row=row, column=5, value=m.get("phone1", ""))
        ws1.cell(row=row, column=6, value=m.get("phone2", ""))
        ws1.cell(row=row, column=7, value=m.get("id", ""))
        row += 1

    for col, w in zip(range(1, 8), [6, 24, 24, 10, 14, 14, 10]):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 2: Payments ──
    ws2 = wb.create_sheet("Payments")
    ws2["A1"] = "RCPM Dues Manager — Payment Records Backup"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:H1")

    headers2 = ["Member SN", "Amount", "Mode", "Date", "Period", "Reference", "Notes", "Recorded At"]
    for col, h in enumerate(headers2, start=1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    row = 4
    for p in sorted(payments, key=lambda x: x.get("date", "")):
        ws2.cell(row=row, column=1, value=p.get("member_sn"))
        ws2.cell(row=row, column=2, value=p.get("amount"))
        ws2.cell(row=row, column=3, value=p.get("mode", ""))
        ws2.cell(row=row, column=4, value=p.get("date", ""))
        ws2.cell(row=row, column=5, value=p.get("period", ""))
        ws2.cell(row=row, column=6, value=p.get("ref", ""))
        ws2.cell(row=row, column=7, value=p.get("notes", ""))
        ws2.cell(row=row, column=8, value=p.get("created_at", ""))
        row += 1

    for col, w in zip(range(1, 9), [10, 12, 12, 12, 10, 16, 24, 20]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: Settings ──
    ws3 = wb.create_sheet("Settings")
    ws3["A1"] = "RCPM Dues Manager — Year Settings Backup"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:B1")

    headers3 = ["Field", "Value"]
    for col, h in enumerate(headers3, start=1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    row = 4
    for k, v in settings.items():
        ws3.cell(row=row, column=1, value=k)
        ws3.cell(row=row, column=2, value=str(v))
        row += 1

    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 45

    # ── Sheet 4: Summary ──
    ws4 = wb.create_sheet("Summary", 0)  # make it first sheet
    ws4["A1"] = "RCPM Dues Manager — Monthly Backup Summary"
    ws4["A1"].font = Font(bold=True, size=16, color="1B3A6B")
    ws4.merge_cells("A1:B1")
    ws4["A3"] = "Backup Generated"
    ws4["B3"] = NOW.strftime("%d %B %Y, %I:%M %p")
    ws4["A4"] = "Total Members"
    ws4["B4"] = len(members)
    ws4["A5"] = "Total Payment Records"
    ws4["B5"] = len(payments)
    ws4["A6"] = "Rotary Year"
    ws4["B6"] = settings.get("year", "—")
    total_paid = sum(float(p.get("amount", 0) or 0) for p in payments)
    ws4["A7"] = "Total Collected (All Records)"
    ws4["B7"] = f"₹{total_paid:,.0f}"
    for r in range(3, 8):
        ws4.cell(row=r, column=1).font = Font(bold=True)
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 30

    path = os.path.join(OUT_DIR, f"RCPM_Backup_{FILE_TAG}.xlsx")
    wb.save(path)
    return path, len(members), len(payments), total_paid


def send_email(json_path, xlsx_path, member_count, payment_count, total_paid):
    msg = MIMEMultipart()
    msg["From"] = GMAIL_ADDRESS
    msg["To"] = ", ".join(BACKUP_RECIPIENTS)
    msg["Subject"] = f"RCPM Dues Manager — Monthly Backup ({MONTH_STR})"

    body = f"""Dear Team,

Please find attached the monthly data backup for RCPM Dues Manager.

📊 BACKUP SUMMARY — {MONTH_STR}
────────────────────────────────
Members on record   : {member_count}
Payment records      : {payment_count}
Total collected (all-time): ₹{total_paid:,.0f}

📎 ATTACHMENTS
- RCPM_Backup_{FILE_TAG}.json  (full raw data, for restore if ever needed)
- RCPM_Backup_{FILE_TAG}.xlsx  (readable spreadsheet — Members, Payments, Settings, Summary)

This is an automated monthly backup. No action is required unless you wish
to store these files separately for your records.

Regards,
RCPM Dues Manager (Automated)
Rotary Club of Patna Millennium
"""
    msg.attach(MIMEText(body, "plain"))

    for path in [json_path, xlsx_path]:
        with open(path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(path)}"')
        msg.attach(part)

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_ADDRESS, BACKUP_RECIPIENTS, msg.as_string())

    print(f"✅ Backup email sent to {', '.join(BACKUP_RECIPIENTS)}")


def main():
    print(f"Starting RCPM backup for {MONTH_STR}...")
    data = fetch_firebase_data()
    json_path = save_json_backup(data)
    print(f"✅ JSON backup saved: {json_path}")
    xlsx_path, mcount, pcount, total_paid = save_excel_backup(data)
    print(f"✅ Excel backup saved: {xlsx_path}")
    send_email(json_path, xlsx_path, mcount, pcount, total_paid)
    print("Done.")


if __name__ == "__main__":
    main()
